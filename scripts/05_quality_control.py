#!/usr/bin/env python3
"""Phase 5: quality control checks for the extraction dataset."""

from __future__ import annotations

import logging
import re
import sys
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd

_BOOTSTRAP_ROOT = Path(__file__).resolve().parents[1]
if str(_BOOTSTRAP_ROOT) not in sys.path:
    sys.path.insert(0, str(_BOOTSTRAP_ROOT))

from scripts._workflow_utils import (
    PROJECT_ROOT,
    ensure_parent_dir,
    load_extraction_dataset,
    resolve_extraction_columns,
)

LOG_PATH = PROJECT_ROOT / "logs" / "05_quality_control.log"
OUTPUT_PATH = PROJECT_ROOT / "data" / "processed" / "quality_control_report.xlsx"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[logging.FileHandler(LOG_PATH), logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

RANGE_RULES: Dict[str, Tuple[float, float]] = {
    "temperature": (-20.0, 300.0),
    "time": (0.1, 10080.0),
    "pressure": (0.1, 100.0),
    "power": (1.0, 10000.0),
    "frequency": (1.0, 200.0),
    "yield": (0.001, 100.0),
    "purity": (0.1, 100.0),
    "tpc": (0.01, 1000.0),
    "tfc": (0.01, 1000.0),
    "ic50": (0.001, 10000.0),
    "ph": (0.0, 14.0),
}

DOI_PATTERN = re.compile(r"^10\.\d{4,9}/\S+$", re.IGNORECASE)


def _flag(row_idx: int, column: str, value: object, issue: str, severity: str) -> Dict[str, object]:
    return {
        "row": int(row_idx),
        "excel_row": int(row_idx) + 4,
        "column": column,
        "value": value,
        "issue": issue,
        "severity": severity,
    }


def check_ranges(df: pd.DataFrame, cols: Dict[str, str]) -> pd.DataFrame:
    flags: List[Dict[str, object]] = []

    for field, (low, high) in RANGE_RULES.items():
        col_name = cols.get(field)
        if not col_name:
            continue

        numeric = pd.to_numeric(df[col_name], errors="coerce")
        mask = numeric.notna() & ((numeric < low) | (numeric > high))

        for idx in df.index[mask]:
            flags.append(
                _flag(
                    idx,
                    col_name,
                    df.at[idx, col_name],
                    f"Outside expected range [{low}, {high}]",
                    "ERROR",
                )
            )

    return pd.DataFrame(flags)


def check_smiles(df: pd.DataFrame) -> pd.DataFrame:
    """Validate SMILES strings using RDKit, reading cells via openpyxl first."""
    flags: List[Dict[str, object]] = []
    try:
        from rdkit import Chem, RDLogger  # pylint: disable=import-error
        from openpyxl import load_workbook  # pylint: disable=import-error
    except ImportError:
        logger.warning("RDKit not installed. Skipping SMILES validation.")
        return pd.DataFrame(flags)

    # Silence repeated RDKit parse warnings; invalid entries are captured as flags below.
    RDLogger.DisableLog("rdApp.error")

    # Read SMILES directly via openpyxl to avoid pandas truncation.
    dataset_path = PROJECT_ROOT / "data" / "processed" / "dataset_with_replacements.xlsx"
    if dataset_path.exists():
        workbook = load_workbook(dataset_path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            smiles_col = None

            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            for col_idx, header in enumerate(header_row, start=1):
                if header and "SMILES" in str(header):
                    smiles_col = col_idx
                    break

            if smiles_col is not None:
                for excel_row, values in enumerate(
                    worksheet.iter_rows(
                        min_row=2,
                        min_col=smiles_col,
                        max_col=smiles_col,
                        values_only=True,
                    ),
                    start=2,
                ):
                    smiles = values[0] if values else None
                    if smiles and str(smiles).strip():
                        if Chem.MolFromSmiles(str(smiles)) is None:
                            flags.append(
                                {
                                    "row": excel_row - 2,
                                    "excel_row": excel_row,
                                    "column": "SMILES",
                                    "value": str(smiles)[:80],
                                    "issue": "Invalid SMILES; RDKit parse failed",
                                    "severity": "ERROR",
                                }
                            )
        finally:
            workbook.close()
        return pd.DataFrame(flags)

    # Fallback to pandas dataframe when replacement dataset is unavailable.
    if "SMILES" in df.columns:
        for idx, smiles in df["SMILES"].items():
            if pd.notna(smiles) and str(smiles).strip():
                if Chem.MolFromSmiles(str(smiles)) is None:
                    flags.append(
                        {
                            "row": int(idx),
                            "excel_row": int(idx) + 4,
                            "column": "SMILES",
                            "value": str(smiles)[:80],
                            "issue": "Invalid SMILES; RDKit parse failed",
                            "severity": "ERROR",
                        }
                    )

    return pd.DataFrame(flags)


def _normalize_doi(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    text = re.sub(r"^https?://(dx\.)?doi\.org/", "", text, flags=re.IGNORECASE)
    return text


def check_dois(df: pd.DataFrame, cols: Dict[str, str]) -> pd.DataFrame:
    flags: List[Dict[str, object]] = []
    doi_col = cols.get("doi")

    if not doi_col:
        return pd.DataFrame(flags)

    for idx, raw_value in df[doi_col].items():
        doi = _normalize_doi(raw_value)
        if not doi:
            continue
        if not DOI_PATTERN.match(doi):
            flags.append(
                _flag(
                    idx,
                    doi_col,
                    str(raw_value)[:100],
                    "Invalid DOI format",
                    "WARNING",
                )
            )

    return pd.DataFrame(flags)


def check_duplicates(df: pd.DataFrame, cols: Dict[str, str]) -> pd.DataFrame:
    flags: List[Dict[str, object]] = []

    dup_fields = ["name", "method", "temperature", "time", "solvent"]
    dup_cols = [cols.get(field) for field in dup_fields if cols.get(field)]

    if len(dup_cols) < 2:
        return pd.DataFrame(flags)

    duplicates = df[df.duplicated(subset=dup_cols, keep=False)]
    for idx in duplicates.index:
        value = " | ".join(str(df.at[idx, col]) for col in dup_cols[:3])
        flags.append(
            _flag(
                idx,
                ", ".join(dup_cols),
                value,
                "Potential duplicate row on key extraction fields",
                "WARNING",
            )
        )

    return pd.DataFrame(flags)


def check_completeness(df: pd.DataFrame) -> pd.DataFrame:
    rows: List[Dict[str, object]] = []
    total = len(df)

    for col in df.columns:
        if col.startswith("_"):
            continue

        non_null = int(df[col].notna().sum())
        missing = total - non_null
        pct = round((100 * non_null / total), 1) if total else 0.0

        rows.append(
            {
                "column": col,
                "filled": non_null,
                "missing": missing,
                "completeness_pct": pct,
            }
        )

    return pd.DataFrame(rows).sort_values("completeness_pct")


def load_best_available_dataset() -> Tuple[pd.DataFrame, Path]:
    processed = PROJECT_ROOT / "data" / "processed" / "dataset_with_replacements.xlsx"
    if processed.exists():
        return pd.read_excel(processed), processed

    raw = PROJECT_ROOT / "data" / "raw" / "Phytochemical_Extraction_10K_Dataset.xlsx"
    return load_extraction_dataset(raw), raw


def main() -> None:
    df, source_path = load_best_available_dataset()
    logger.info("Loaded %s rows from %s", len(df), source_path)

    cols = resolve_extraction_columns(df.columns)

    range_flags = check_ranges(df, cols)
    logger.info("Range issues: %s", len(range_flags))

    smiles_flags = check_smiles(df)
    logger.info("SMILES issues: %s", len(smiles_flags))

    doi_flags = check_dois(df, cols)
    logger.info("DOI issues: %s", len(doi_flags))

    duplicate_flags = check_duplicates(df, cols)
    logger.info("Duplicate warnings: %s", len(duplicate_flags))

    completeness = check_completeness(df)

    all_flags = pd.concat(
        [range_flags, smiles_flags, doi_flags, duplicate_flags],
        ignore_index=True,
    )

    if all_flags.empty:
        all_flags = pd.DataFrame(columns=["row", "excel_row", "column", "value", "issue", "severity"])

    summary = pd.DataFrame(
        [
            {
                "metric": "total_rows",
                "value": len(df),
            },
            {
                "metric": "total_issues",
                "value": len(all_flags),
            },
            {
                "metric": "errors",
                "value": int((all_flags["severity"] == "ERROR").sum()) if not all_flags.empty else 0,
            },
            {
                "metric": "warnings",
                "value": int((all_flags["severity"] == "WARNING").sum()) if not all_flags.empty else 0,
            },
            {
                "metric": "source_path",
                "value": str(source_path),
            },
        ]
    )

    ensure_parent_dir(OUTPUT_PATH)
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        all_flags.to_excel(writer, sheet_name="Issues", index=False)
        completeness.to_excel(writer, sheet_name="Completeness", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)

    logger.info("=" * 60)
    logger.info("QUALITY CONTROL COMPLETE")
    logger.info("Issues: %s", len(all_flags))
    logger.info("Errors: %s", int((all_flags["severity"] == "ERROR").sum()) if not all_flags.empty else 0)
    logger.info("Warnings: %s", int((all_flags["severity"] == "WARNING").sum()) if not all_flags.empty else 0)
    logger.info("Report: %s", OUTPUT_PATH)
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
