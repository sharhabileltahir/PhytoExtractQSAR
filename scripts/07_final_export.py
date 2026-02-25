#!/usr/bin/env python3
"""Phase 7: assemble final publication-ready workbook from template + processed data."""

from __future__ import annotations

import logging
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook

_BOOTSTRAP_ROOT = Path(__file__).resolve().parents[1]
if str(_BOOTSTRAP_ROOT) not in sys.path:
    sys.path.insert(0, str(_BOOTSTRAP_ROOT))

from scripts._workflow_utils import (
    PROJECT_ROOT,
    load_extraction_dataset,
    normalize_label,
    resolve_extraction_columns,
)

LOG_PATH = PROJECT_ROOT / "logs" / "07_final_export.log"
TEMPLATE_PATH = PROJECT_ROOT / "data" / "raw" / "Phytochemical_Extraction_10K_Dataset.xlsx"
REPLACED_PATH = PROJECT_ROOT / "data" / "processed" / "dataset_with_replacements.xlsx"
DESCRIPTOR_PATH = PROJECT_ROOT / "data" / "processed" / "molecular_descriptors_rdkit.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "output"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[logging.FileHandler(LOG_PATH), logging.StreamHandler()],
)
logger = logging.getLogger(__name__)


def to_native(value: object) -> object:
    if pd.isna(value):
        return None

    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()

    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return value

    return value


def build_header_index(worksheet, header_row: int) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for col_idx in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=header_row, column=col_idx).value
        if value is None:
            continue
        normalized = normalize_label(value)
        if normalized and normalized not in header_map:
            header_map[normalized] = col_idx
    return header_map


def map_dataframe_columns(df: pd.DataFrame, header_index: Dict[str, int]) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for column in df.columns:
        normalized = normalize_label(column)
        if normalized in header_index:
            mapping[column] = header_index[normalized]
    return mapping


def clear_sheet_rows(worksheet, start_row: int, col_indices: List[int], target_rows: int) -> None:
    last_row = max(worksheet.max_row, start_row + max(target_rows, 1) - 1)
    for row_idx in range(start_row, last_row + 1):
        for col_idx in col_indices:
            worksheet.cell(row=row_idx, column=col_idx).value = None


def write_dataframe_to_sheet(df: pd.DataFrame, worksheet, header_row: int, start_row: int) -> Tuple[int, int]:
    header_index = build_header_index(worksheet, header_row)
    mapping = map_dataframe_columns(df, header_index)

    if not mapping:
        raise RuntimeError(f"No matching columns found for sheet '{worksheet.title}'")

    col_indices = sorted(set(mapping.values()))
    clear_sheet_rows(worksheet, start_row, col_indices, len(df))

    for row_offset, (_, row) in enumerate(df.iterrows()):
        row_idx = start_row + row_offset
        for df_col, col_idx in mapping.items():
            value = to_native(row[df_col])
            if value is not None:
                worksheet.cell(row=row_idx, column=col_idx).value = value

    return len(df), len(mapping)


def load_best_dataset() -> pd.DataFrame:
    if REPLACED_PATH.exists():
        return pd.read_excel(REPLACED_PATH)

    return load_extraction_dataset(TEMPLATE_PATH)


def build_report(df_full: pd.DataFrame, final_path: Path) -> str:
    cols = resolve_extraction_columns(df_full.columns)

    name_col = cols.get("name")
    plant_col = cols.get("plant_source")
    method_col = cols.get("method")

    unique_compounds = int(df_full[name_col].nunique()) if name_col else 0
    unique_plants = int(df_full[plant_col].nunique()) if plant_col else 0
    unique_methods = int(df_full[method_col].nunique()) if method_col else 0

    if "_data_source" in df_full.columns:
        source_counts = df_full["_data_source"].fillna("unknown").value_counts().to_string()
    else:
        source_counts = f"synthetic    {len(df_full)}"

    timestamp = datetime.now().strftime("%Y%m%d")

    report = (
        "=" * 60
        + f"\nFINAL DATASET REPORT - {timestamp}\n"
        + "=" * 60
        + f"\nTotal records:           {len(df_full)}"
        + f"\nUnique compounds:        {unique_compounds}"
        + f"\nUnique plant sources:    {unique_plants}"
        + f"\nExtraction methods:      {unique_methods}"
        + "\n\nDATA PROVENANCE:\n"
        + source_counts
        + f"\n\nOutput: {final_path}"
        + "\n" + "=" * 60 + "\n"
        + "\nMANUSCRIPT CHECKLIST:\n"
        + "[ ] All DOIs verified and accessible\n"
        + "[ ] SMILES validated against PubChem\n"
        + "[ ] Molecular descriptors recomputed via RDKit\n"
        + "[ ] Extraction parameters pass quality-control ranges\n"
        + "[ ] Quality-control report reviewed\n"
        + "[ ] Data dictionary updated\n"
        + "[ ] Supplementary materials prepared\n"
        + "=" * 60
    )
    return report


def main() -> None:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template workbook not found: {TEMPLATE_PATH}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    df_full = load_best_dataset()
    tracking_cols = [col for col in df_full.columns if str(col).startswith("_")]
    df_clean = df_full.drop(columns=tracking_cols, errors="ignore")

    desc_df = pd.read_excel(DESCRIPTOR_PATH) if DESCRIPTOR_PATH.exists() else None

    wb = load_workbook(TEMPLATE_PATH)

    extraction_sheet = wb["Extraction_Data"]
    rows_written, cols_mapped = write_dataframe_to_sheet(
        df_clean,
        extraction_sheet,
        header_row=3,
        start_row=4,
    )
    logger.info("Extraction sheet updated: %s rows, %s mapped columns", rows_written, cols_mapped)

    if desc_df is not None and not desc_df.empty and "Molecular_Descriptors" in wb.sheetnames:
        descriptor_sheet = wb["Molecular_Descriptors"]
        d_rows, d_cols = write_dataframe_to_sheet(
            desc_df,
            descriptor_sheet,
            header_row=2,
            start_row=3,
        )
        logger.info("Descriptor sheet updated: %s rows, %s mapped columns", d_rows, d_cols)
    else:
        logger.info("Descriptor sheet unchanged (no computed descriptor file found)")

    timestamp = datetime.now().strftime("%Y%m%d")
    final_path = OUTPUT_DIR / f"Phytochemical_Extraction_Dataset_FINAL_{timestamp}.xlsx"
    wb.save(final_path)

    report_text = build_report(df_full, final_path)
    report_path = OUTPUT_DIR / f"dataset_report_{timestamp}.txt"
    report_path.write_text(report_text, encoding="utf-8")

    logger.info("=" * 60)
    logger.info("FINAL EXPORT COMPLETE")
    logger.info("Workbook: %s", final_path)
    logger.info("Report: %s", report_path)
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
